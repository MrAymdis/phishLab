"""报表中心服务：概览指标、单演练报表、部门/趋势/个人报表、异步导出。

数据来源：campaign_target 事实表直查（按投递时间窗口，分子分母同窗口）；
stat_daily 日归档由 stat_aggregate 任务每日聚合（平台/部门/场景/用户四维度），
个人画像趋势读用户维度序列；所有比率由数据库计数实时计算，禁止硬编码；
空表返回零值结构，不抛异常。
"""
from datetime import datetime, timedelta
from uuid import uuid4

from sqlalchemy import case, func, or_, select

from app.core.audit import record_audit
from app.core.deps import apply_data_scope
from app.core.errors import BizError, ErrorCode
from app.modules.analytics.models import StatDaily
from app.modules.campaign.models import Campaign, CampaignStat, CampaignTarget
from app.modules.org.models import EmpDept, EmpRiskProfile, EmpUser
from app.modules.template.models import AttachmentPayload, EmailTemplate, LandingPage, QrAsset
from app.modules.tracking.models import TrackEvent
from app.modules.training.models import Course, TrainingAssignment

_DAYS = {"7d": 7, "month": 30, "quarter": 90, "year": 365}
_TYPE_CN = {"mail": "邮件", "sms": "短信", "social": "社交媒体", "usb": "USB"}


# ---------- 公共小工具 ----------

def _window_days(range_: str) -> int:
    return _DAYS.get(range_, 30)


def _pct(n, d) -> float:
    """比率（保留 1 位小数），分母为 0 返回 0.0。MySQL SUM 返回 Decimal，统一转 float。"""
    n, d = float(n or 0), float(d or 0)
    return round(n / d * 100, 1) if d else 0.0


def _fmt_rate(n, d) -> str:
    n, d = float(n or 0), float(d or 0)
    return f"{n / d * 100:.1f}" if d else "0.0"


def _bar(n) -> int:
    """横向条形图长度：计数 *25，封顶 100、保底 5。"""
    return max(5, min(int(n * 25), 100))


_SENT_STATUS = ("sent", "delivered", "bounced", "failed")


def _agg_platform(db, since):
    """窗口内投递聚合：campaign_target 事实表直查（按投递时间窗口）。

    实时口径以事实表为准（stat_daily 为日归档，用于留存清理后的长期回放）；
    所有计数限定在「投递时间 >= since」的记录内，保证分子分母同窗口。
    """
    sent = CampaignTarget.send_status.in_(_SENT_STATUS)
    base = (
        select(
            func.count(CampaignTarget.id),  # 投递人次（实际发出）
            func.sum(case((CampaignTarget.send_status.in_(("sent", "delivered")), 1), else_=0)),  # 送达
            func.sum(case((CampaignTarget.first_open_at.is_not(None), 1), else_=0)),  # 打开
            func.sum(case((CampaignTarget.first_click_at.is_not(None), 1), else_=0)),  # 点击
            func.sum(case((CampaignTarget.submit_flag == 1, 1), else_=0)),  # 提交
            func.sum(case((CampaignTarget.report_flag == 1, 1), else_=0)),  # 举报
        )
        .where(sent)
    )
    row = db.execute(base.where(CampaignTarget.sent_at >= since)).one()
    if not any(row):  # 窗口内无投递 → 全量回退
        row = db.execute(base).one()
    return [0, *[int(v or 0) for v in row]]  # 首项 campaign_cnt 占位，由调用方覆盖


def _bucket_ts(buckets, dt, since, n_buckets, week_bucket):
    """将事件时间戳落入窗口分桶（week_bucket 按周且尾桶收口，否则按天）。

    30 天窗口 / 7 天桶 = 4.28 桶，最后的几天会落入第 5 个不存在的桶，
    故按周时对越界索引做 min 截断并入尾桶（与按天分桶语义一致）。
    """
    if dt is None:
        return
    idx = (dt.date() - since).days // 7 if week_bucket else (dt.date() - since).days
    if week_bucket:
        idx = min(idx, n_buckets - 1)
    if not 0 <= idx < n_buckets:
        return
    buckets[idx] += 1


def _campaign_stat_sums(db):
    """campaign_stat 汇总回退（stat_daily 为空时）。"""
    row = db.execute(
        select(
            func.coalesce(func.sum(CampaignStat.delivered_cnt), 0),
            func.coalesce(func.sum(CampaignStat.open_cnt), 0),
            func.coalesce(func.sum(CampaignStat.click_cnt), 0),
            func.coalesce(func.sum(CampaignStat.submit_cnt), 0),
            func.coalesce(func.sum(CampaignStat.report_cnt), 0),
        )
    ).one()
    return [int(v or 0) for v in row]


def _parse_ua(ua: str) -> str:
    """从 UA 提取「浏览器 · 系统」，用于指纹饼图降级。"""
    browser = "Other"
    if "Edg/" in ua:
        browser = "Edge"
    elif "Chrome/" in ua:
        browser = "Chrome"
    elif "Firefox/" in ua:
        browser = "Firefox"
    elif "Safari/" in ua:
        browser = "Safari"
    os = "Other"
    if "Windows" in ua:
        os = "Windows"
    elif "iPhone" in ua or "iPad" in ua or "iOS" in ua:
        os = "iOS"
    elif "Mac OS" in ua or "Macintosh" in ua:
        os = "macOS"
    elif "Android" in ua:
        os = "Android"
    elif "Linux" in ua:
        os = "Linux"
    return f"{browser} · {os}"


# ---------- 数据概览 ----------

def overview_metrics(db, account, range_: str) -> dict:
    """数据概览：核心指标 + 方式分布 + 趋势 + TOP5 + 指纹饼图，随 range 联动。

    平台级聚合指标面向管理员大盘，不按部门数据权限过滤。
    """
    days = _window_days(range_)
    window_start = datetime.now() - timedelta(days=days)
    since = window_start.date()

    # 平台聚合：窗口 → 全量 → campaign_stat 回退
    campaign_cnt, target, delivered, open_cnt, click_cnt, submit, report = _agg_platform(db, since)
    if not (campaign_cnt or target or delivered):
        delivered, open_cnt, click_cnt, submit, report = _campaign_stat_sums(db)
        target = int(db.scalar(select(func.coalesce(func.sum(Campaign.target_count), 0))) or 0)

    # 演练次数：窗口内有则取窗口内，否则取总量
    in_win = int(db.scalar(
        select(func.count()).select_from(Campaign).where(Campaign.created_at >= window_start)
    ) or 0)
    campaign_cnt = in_win or int(db.scalar(select(func.count()).select_from(Campaign)) or 0)

    # 培训通过率（按学习任务计算）
    completed = int(db.scalar(
        select(func.count()).select_from(TrainingAssignment).where(TrainingAssignment.status == "completed")
    ) or 0)
    assigned = int(db.scalar(select(func.count()).select_from(TrainingAssignment)) or 0)
    training_rate = _fmt_rate(completed, assigned)

    high_risk = int(db.scalar(
        select(func.count()).select_from(EmpRiskProfile).where(EmpRiskProfile.risk_level >= 3)
    ) or 0)

    core_metrics = [
        {"title": "演练次数", "value": campaign_cnt, "suffix": " 场", "accent": "blue"},
        {"title": "演练人数", "value": f"{target:,}", "suffix": " 人", "accent": "teal"},
        {"title": "平均中招率", "value": _fmt_rate(submit, target), "suffix": " %", "accent": "orange"},
        {"title": "平均举报率", "value": _fmt_rate(report, target), "suffix": " %", "accent": "green"},
        {"title": "培训通过率", "value": training_rate, "suffix": " %", "accent": "purple"},
        {"title": "高危人员数", "value": high_risk, "suffix": " 人", "accent": "red"},
    ]

    # TOP5 高危人员（历史中招次数）
    top_rows = db.execute(
        select(EmpRiskProfile.phish_count, EmpUser.name, EmpDept.name)
        .join(EmpUser, EmpUser.id == EmpRiskProfile.user_id)
        .join(EmpDept, EmpDept.id == EmpUser.dept_id, isouter=True)
        .order_by(EmpRiskProfile.phish_count.desc())
        .limit(5)
    ).all()
    top_persons = [
        {"name": name or "", "dept": dept or "", "count": f"{int(n or 0)}次", "bar": _bar(n or 0)}
        for n, name, dept in top_rows
    ]

    # TOP5 中招部门（真实提交事件按部门聚合，同窗口）
    dept_stmt = (
        select(EmpDept.name, func.count(CampaignTarget.id))
        .join(EmpUser, EmpUser.id == CampaignTarget.user_id)
        .join(EmpDept, EmpDept.id == EmpUser.dept_id, isouter=True)
        .where(CampaignTarget.submit_flag == 1, CampaignTarget.sent_at >= since)
        .group_by(EmpDept.name)
        .order_by(func.count(CampaignTarget.id).desc())
        .limit(5)
    )
    top_depts = [
        {"name": name or "未知部门", "count": f"{int(n or 0)}次", "bar": _bar(n or 0)}
        for name, n in db.execute(dept_stmt).all()
    ]

    # 进行中演练实时投递漏斗
    running_camps = db.scalars(
        select(Campaign)
        .where(Campaign.status.in_(("sending", "running")))
        .order_by(Campaign.id.desc())
    ).all()
    running_ids = [c.id for c in running_camps]
    live_campaign_name = running_camps[0].name if running_camps else ""
    live = [0] * 5
    if running_ids:
        row = db.execute(
            select(
                func.coalesce(func.sum(CampaignStat.delivered_cnt), 0),
                func.coalesce(func.sum(CampaignStat.open_cnt), 0),
                func.coalesce(func.sum(CampaignStat.click_cnt), 0),
                func.coalesce(func.sum(CampaignStat.submit_cnt), 0),
                func.coalesce(func.sum(CampaignStat.report_cnt), 0),
            ).where(CampaignStat.campaign_id.in_(running_ids))
        ).one()
        live = [int(v or 0) for v in row]
    live_defs = [
        ("已投递", live[0], "#378ADD"),
        ("已阅读", live[1], "#13C2C2"),
        ("已点击", live[2], "#FAAD14"),
        ("已提交", live[3], "#A32D2D"),
        ("已举报", live[4], "#7F77DD"),
    ]
    live_stats = []
    for label, val, color in live_defs:
        bar = 100 if label == "已投递" else max(0, min(round(val / live[0] * 100) if live[0] else 0, 100))
        live_stats.append({"label": label, "value": f"{val:,}", "bar": bar, "color": color})

    # 素材运营数据
    ops_data = [
        {"label": "邮件钓鱼模板", "value": int(db.scalar(select(func.count()).select_from(EmailTemplate)) or 0)},
        {"label": "口令钓鱼模板", "value": int(db.scalar(select(func.count()).select_from(LandingPage)) or 0)},
        {"label": "二维码钓鱼模板", "value": int(db.scalar(select(func.count()).select_from(QrAsset)) or 0)},
        {"label": "附件载荷", "value": int(db.scalar(select(func.count()).select_from(AttachmentPayload)) or 0)},
        {"label": "落地页", "value": int(db.scalar(select(func.count()).select_from(LandingPage)) or 0)},
    ]

    # 待开始/筹备中的演练计划
    plans_rows = db.scalars(
        select(Campaign)
        .where(Campaign.status.in_(("draft", "scheduled")))
        .order_by(Campaign.schedule_at.is_(None), Campaign.schedule_at.asc(), Campaign.id.desc())  # 无排期的排最后（兼容 MySQL）
        .limit(5)
    ).all()
    plans = []
    for c in plans_rows:
        d = c.schedule_at or c.started_at
        plans.append({
            "name": c.name,
            "date": d.strftime("%m-%d") if d else "",
            "type": _TYPE_CN.get(c.type, c.type or "") + "钓鱼演练",
            "target": f"{int(c.target_count or 0)}人",
            "status": {"scheduled": "待开始", "draft": "筹备中"}.get(c.status, "进行中"),
        })

    # 演练方式分布
    type_rows = db.execute(select(Campaign.type, func.count()).group_by(Campaign.type)).all()
    channel_dist = [
        {"name": f"{_TYPE_CN.get(t, t or '其他')}钓鱼演练", "value": int(n or 0)}
        for t, n in type_rows
    ]

    # 中招趋势（7d 按天；month/quarter 按周）
    if range_ == "7d":
        n_buckets, week_bucket = 7, False
        labels = [(window_start + timedelta(days=i)).strftime("%m-%d") for i in range(7)]
    else:
        n_buckets, week_bucket = 4 if range_ == "month" else 12, True
        labels = [f"W{i + 1}" for i in range(n_buckets)]
    victims, bucket_target = [0] * n_buckets, [0] * n_buckets
    for (t,) in db.execute(
        select(CampaignTarget.submit_at)
        .where(CampaignTarget.submit_flag == 1, CampaignTarget.sent_at >= since)
    ).all():
        _bucket_ts(victims, t, since, n_buckets, week_bucket)
    for (t,) in db.execute(
        select(CampaignTarget.sent_at)
        .where(CampaignTarget.send_status.in_(_SENT_STATUS), CampaignTarget.sent_at >= since)
    ).all():
        _bucket_ts(bucket_target, t, since, n_buckets, week_bucket)
    victim_rates = [_pct(victims[i], bucket_target[i]) for i in range(n_buckets)]

    trend = {"labels": labels, "victims": victims, "victimRates": victim_rates}

    # 指纹饼图：fp_hash 无法还原浏览器，直接从 track_event.ua 聚合
    ua_rows = db.scalars(
        select(TrackEvent.ua)
        .where(TrackEvent.ua.is_not(None), TrackEvent.ua != "")
        .limit(500)
    ).all()
    ua_counts: dict[str, int] = {}
    for ua in ua_rows:
        name = _parse_ua(ua)
        ua_counts[name] = ua_counts.get(name, 0) + 1
    top_ua = sorted(ua_counts.items(), key=lambda kv: -kv[1])[:6]
    fingerprints = [{"name": name, "value": cnt} for name, cnt in top_ua]

    return {
        "coreMetrics": core_metrics,
        "topPersons": top_persons,
        "topDepts": top_depts,
        "liveStats": live_stats,
        "liveCampaignName": live_campaign_name,
        "opsData": ops_data,
        "plans": plans,
        "channelDist": channel_dist,
        "trend": trend,
        "fingerprints": fingerprints,
    }


# ---------- 单次演练报表 ----------

def campaign_report(db, account, campaign_id: int) -> dict:
    """单次演练报表：指标卡 + 漏斗（逐级转化率）+ 中招明细 + 日趋势。"""
    campaign = db.get(Campaign, campaign_id)
    if campaign is None:
        raise BizError(ErrorCode.NOT_FOUND)

    target = int(campaign.target_count or 0)
    stat = db.get(CampaignStat, campaign_id)
    delivered = int(stat.delivered_cnt or 0) if stat else 0
    open_cnt = int(stat.open_cnt or 0) if stat else 0
    click_cnt = int(stat.click_cnt or 0) if stat else 0
    submit_cnt = int(stat.submit_cnt or 0) if stat else 0
    report_cnt = int(stat.report_cnt or 0) if stat else 0
    attach_cnt = int(db.scalar(
        select(func.count()).select_from(TrackEvent)
        .where(TrackEvent.campaign_id == campaign_id, TrackEvent.event_type == "attach_run")
    ) or 0)

    submit_rate = _pct(submit_cnt, target)
    metrics = [
        {"title": "发送数", "value": target, "suffix": " 封", "sub": "目标规模", "accent": "blue"},
        {"title": "打开数", "value": open_cnt, "suffix": " 封", "sub": f"打开率 {_pct(open_cnt, target):.1f}%", "accent": "teal"},
        {"title": "点击数", "value": click_cnt, "suffix": " 次", "sub": f"点击率 {_pct(click_cnt, target):.1f}%", "accent": "orange"},
        {"title": "中招数", "value": submit_cnt, "suffix": " 人", "sub": f"中招率 {submit_rate:.1f}%", "accent": "red"},
        {"title": "举报数", "value": report_cnt, "suffix": " 封", "sub": f"举报率 {_pct(report_cnt, target):.1f}%", "accent": "green"},
        {"title": "综合得分", "value": round(100 - max(submit_rate * 2, 0)), "suffix": " 分", "accent": "purple"},
    ]

    # 漏斗：逐级转化率（相对上一阶段）
    sent = delivered or target
    funnel = [
        {"name": "发送成功", "value": sent, "rate": _pct(sent, target)},
        {"name": "已阅读", "value": open_cnt, "rate": _pct(open_cnt, target)},
        {"name": "已点击", "value": click_cnt, "rate": _pct(click_cnt, open_cnt)},
        {"name": "输入数据", "value": submit_cnt, "rate": _pct(submit_cnt, click_cnt)},
        {"name": "已举报", "value": report_cnt, "rate": _pct(report_cnt, submit_cnt)},
        {"name": "附件运行", "value": attach_cnt, "rate": _pct(attach_cnt, report_cnt)},
    ]

    # 中招明细（提交过或点击过）
    victim_rows = db.execute(
        select(CampaignTarget, EmpUser.name, EmpUser.email, EmpDept.name)
        .join(EmpUser, EmpUser.id == CampaignTarget.user_id, isouter=True)
        .join(EmpDept, EmpDept.id == EmpUser.dept_id, isouter=True)
        .where(
            CampaignTarget.campaign_id == campaign_id,
            or_(CampaignTarget.submit_flag == 1, CampaignTarget.click_count > 0),
        )
        .order_by(CampaignTarget.submit_at.desc(), CampaignTarget.first_click_at.desc())
        .limit(100)
    ).all()
    victims = [
        {
            "name": name or "", "dept": dept or "", "email": email or "",
            "clicks": int(t.click_count or 0), "input_pwd": bool(t.submit_flag),
            "first_open": t.first_open_at.strftime("%Y-%m-%d %H:%M") if t.first_open_at else "",
        }
        for t, name, email, dept in victim_rows
    ]

    # 日趋势（track_event 按天聚合）
    date_col = func.date(TrackEvent.created_at)
    trend_rows = db.execute(
        select(TrackEvent.event_type, date_col, func.count())
        .where(TrackEvent.campaign_id == campaign_id, TrackEvent.event_type.in_(("open", "click", "submit")))
        .group_by(TrackEvent.event_type, date_col)
        .order_by(date_col)
    ).all()
    days = sorted({r[1] for r in trend_rows})
    series = {d: {"open": 0, "click": 0, "submit": 0} for d in days}
    for et, d, n in trend_rows:
        series[d][et] = int(n or 0)
    daily_trend = {
        "labels": [f"D{i + 1}" for i in range(len(days))],
        "opens": [series[d]["open"] for d in days],
        "clicks": [series[d]["click"] for d in days],
        "submits": [series[d]["submit"] for d in days],
    }

    # 部门对比明细（该演练内按部门聚合：投递/中招/举报）
    dept_stmt = (
        select(EmpDept.name,
               func.sum(case((CampaignTarget.send_status.in_(_SENT_STATUS), 1), else_=0)),
               func.sum(case((CampaignTarget.submit_flag == 1, 1), else_=0)),
               func.sum(case((CampaignTarget.report_flag == 1, 1), else_=0)))
        .join(EmpUser, EmpUser.id == CampaignTarget.user_id)
        .join(EmpDept, EmpDept.id == EmpUser.dept_id, isouter=True)
        .where(CampaignTarget.campaign_id == campaign_id)
        .group_by(EmpDept.name)
        .order_by(func.sum(case((CampaignTarget.submit_flag == 1, 1), else_=0)).desc())
    )
    dept_compare = [
        {"dept": name or "未知部门", "sent": int(s or 0), "victim": int(v or 0),
         "victimRate": _pct(v or 0, s or 0), "report": int(r or 0),
         "reportRate": _pct(r or 0, s or 0)}
        for name, s, v, r in db.execute(dept_stmt).all()
    ]

    return {"metrics": metrics, "funnel": funnel, "victims": victims, "dailyTrend": daily_trend,
            "deptCompare": dept_compare}


# ---------- 部门 / 趋势 / 个人报表 ----------

def department_report(db, account, range_: str) -> dict:
    """部门横向对比（campaign_target 直查，按部门数据权限过滤，投递时间窗口）。"""
    days = _window_days(range_)
    since = (datetime.now() - timedelta(days=days)).date()

    submit_c = func.sum(case((CampaignTarget.submit_flag == 1, 1), else_=0))
    stmt = (
        select(
            EmpUser.dept_id,
            func.sum(case((CampaignTarget.send_status.in_(_SENT_STATUS), 1), else_=0)),  # 投递人次
            func.sum(case((CampaignTarget.first_open_at.is_not(None), 1), else_=0)),
            func.sum(case((CampaignTarget.first_click_at.is_not(None), 1), else_=0)),
            submit_c,
            func.sum(case((CampaignTarget.report_flag == 1, 1), else_=0)),
            EmpDept.name,
        )
        .join(EmpUser, EmpUser.id == CampaignTarget.user_id)
        .join(EmpDept, EmpDept.id == EmpUser.dept_id, isouter=True)
        .group_by(EmpUser.dept_id, EmpDept.name)
        .order_by(submit_c.desc())
    )
    stmt = apply_data_scope(db, stmt, account, dept_col=EmpUser.dept_id)
    rows = db.execute(stmt.where(CampaignTarget.sent_at >= since)).all()
    if not rows:  # 窗口内无投递回退全量
        rows = db.execute(stmt).all()

    # 部门在职人数 + 培训完成率（EmpUser / TrainingAssignment 辅助聚合）
    total_map = dict(db.execute(
        select(EmpUser.dept_id, func.count(EmpUser.id)).where(EmpUser.status == 1).group_by(EmpUser.dept_id)
    ).all())
    train_map = dict(db.execute(
        select(EmpUser.dept_id,
               func.sum(case((TrainingAssignment.status == "completed", 1), else_=0)),
               func.count(TrainingAssignment.id))
        .join(TrainingAssignment, TrainingAssignment.user_id == EmpUser.id)
        .group_by(EmpUser.dept_id)
    ).all())

    rows_out, labels, submit_rates = [], [], []
    for dim_id, target, open_cnt, click_cnt, submit_cnt, report_cnt, name in rows:
        target = float(target or 0)
        label = name or f"部门{dim_id or ''}"
        sr = _pct(submit_cnt or 0, target)
        completed, assigned = train_map.get(dim_id, (0, 0))
        rows_out.append({
            "dept": label, "targetCount": int(target),  # 覆盖次数（投递人次）
            "victim": int(submit_cnt or 0), "report": int(report_cnt or 0),
            "total": int(total_map.get(dim_id, 0) or 0),  # 部门在职人数
            "trainRate": _pct(completed or 0, assigned or 0),
            "openRate": _pct(open_cnt or 0, target),
            "clickRate": _pct(click_cnt or 0, target),
            "submitRate": sr,
            "reportRate": _pct(report_cnt or 0, target),
        })
        labels.append(label)
        submit_rates.append(sr)
    return {"rows": rows_out, "labels": labels, "submitRates": submit_rates}


def dept_persons_report(db, account, dept_id: int, range_: str) -> dict:
    """部门内人员明细：参与次数/中招率/风险等级（campaign_target 直查，过部门数据权限）。"""
    days = _window_days(range_)
    since = (datetime.now() - timedelta(days=days)).date()

    submit_c = func.sum(case((CampaignTarget.submit_flag == 1, 1), else_=0))
    stmt = (
        select(EmpUser.id, EmpUser.name, EmpUser.emp_no, EmpUser.position, EmpDept.name,
               func.count(CampaignTarget.id), submit_c, EmpRiskProfile.risk_level)
        .join(CampaignTarget, CampaignTarget.user_id == EmpUser.id)
        .join(EmpDept, EmpDept.id == EmpUser.dept_id, isouter=True)
        .join(EmpRiskProfile, EmpRiskProfile.user_id == EmpUser.id, isouter=True)
        .where(EmpUser.dept_id == dept_id, CampaignTarget.send_status.in_(_SENT_STATUS))
        .group_by(EmpUser.id, EmpUser.name, EmpUser.emp_no, EmpUser.position, EmpDept.name, EmpRiskProfile.risk_level)
        .order_by(submit_c.desc(), EmpUser.id)
    )
    stmt = apply_data_scope(db, stmt, account, dept_col=EmpUser.dept_id)
    rows = db.execute(stmt.where(CampaignTarget.sent_at >= since)).all()
    if not rows:  # 窗口内无投递回退全量
        rows = db.execute(stmt).all()
    return {
        "rows": [
            {"id": uid, "name": name or "", "empNo": emp_no or "", "dept": dept or "",
             "position": position or "", "drills": int(n or 0),
             "victimRate": _pct(v or 0, n or 0),
             "risk": {1: "low", 2: "mid", 3: "high"}.get(risk_level, "low")}
            for uid, name, emp_no, position, dept, n, v, risk_level in rows
        ]
    }


def trend_report(db, account, range_: str) -> dict:
    """跨演练月度趋势 + 场景维度（campaign_target 直查，按投递时间窗口）。平台级，不过数据权限。"""
    days = _window_days(range_)
    since = (datetime.now() - timedelta(days=days)).date()
    months = {"7d": 1, "month": 3, "quarter": 6, "year": 12}.get(range_, 3)

    # 最近 N 个月标签
    now = datetime.now()
    month_keys = []
    for i in range(months - 1, -1, -1):
        y, m = now.year, now.month - i
        while m <= 0:
            m += 12
            y -= 1
        month_keys.append((y, m))
    labels = [f"{m}月" for _, m in month_keys]

    # 投递按月聚合（窗口内，分子分母同窗口；提交/举报按投递月份归属）
    year_expr, month_expr = func.extract("year", CampaignTarget.sent_at), func.extract("month", CampaignTarget.sent_at)
    month_rows = db.execute(
        select(year_expr, month_expr, func.count(CampaignTarget.id),
               func.sum(case((CampaignTarget.submit_flag == 1, 1), else_=0)),
               func.sum(case((CampaignTarget.report_flag == 1, 1), else_=0)))
        .where(CampaignTarget.send_status.in_(_SENT_STATUS), CampaignTarget.sent_at >= since)
        .group_by(year_expr, month_expr)
    ).all()
    by_month = {f"{int(y)}-{int(m):02d}": (t or 0, s or 0, r or 0) for y, m, t, s, r in month_rows}

    # 演练数按月（窗口内）
    c_rows = db.execute(
        select(func.extract("year", Campaign.created_at), func.extract("month", Campaign.created_at), func.count())
        .where(Campaign.created_at >= datetime(since.year, since.month, since.day))
        .group_by(func.extract("year", Campaign.created_at), func.extract("month", Campaign.created_at))
    ).all()
    c_by_month = {f"{int(y)}-{int(m):02d}": int(n or 0) for y, m, n in c_rows}

    campaign_counts, submit_rates, report_rates = [], [], []
    for y, m in month_keys:
        key = f"{y}-{m:02d}"
        t, s, r = by_month.get(key, (0, 0, 0))
        campaign_counts.append(c_by_month.get(key, 0))
        submit_rates.append(_pct(s, t))
        report_rates.append(_pct(r, t))

    # 场景维度（按演练方式，窗口内 → 全量回退）
    scene_stmt = (
        select(Campaign.type, func.count(CampaignTarget.id),
               func.sum(case((CampaignTarget.submit_flag == 1, 1), else_=0)),
               func.sum(case((CampaignTarget.report_flag == 1, 1), else_=0)))
        .join(Campaign, Campaign.id == CampaignTarget.campaign_id)
        .where(CampaignTarget.send_status.in_(_SENT_STATUS))
        .group_by(Campaign.type)
    )
    s_rows = db.execute(scene_stmt.where(CampaignTarget.sent_at >= since)).all()
    if not s_rows:
        s_rows = db.execute(scene_stmt).all()
    scenes = [
        {"scene": _TYPE_CN.get(t, t or "其他") + "钓鱼演练", "targetCount": int(n or 0),
         "submitRate": _pct(s or 0, n or 0), "reportRate": _pct(r or 0, n or 0)}
        for t, n, s, r in s_rows
    ]

    return {"labels": labels, "campaignCounts": campaign_counts,
            "submitRates": submit_rates, "reportRates": report_rates, "scenes": scenes}


def _risk_color(v) -> str:
    return "#F5222D" if v >= 71 else ("#FAAD14" if v >= 31 else "#52C41A")


def _level_of(total: int) -> str:
    return "低" if total <= 30 else ("中" if total <= 70 else "高")


def personal_report(db, account, user_id: int) -> dict:
    """员工个人安全档案：五维雷达 + 风险趋势 + 行为时间轴 + 培训记录。"""
    user = db.get(EmpUser, user_id)
    if user is None:
        raise BizError(ErrorCode.NOT_FOUND)

    profile = db.get(EmpRiskProfile, user_id)
    total = int(profile.total_score) if profile else int(user.initial_risk or 50)
    dim_defs = [
        ("邮件识别", int(profile.email_recognize) if profile else 50),
        ("链接点击", int(profile.link_click) if profile else 50),
        ("密码提交", int(profile.pwd_submit) if profile else 50),
        ("附件下载", int(profile.attach_run) if profile else 50),
        ("举报意识", int(profile.report_awareness) if profile else 50),
    ]
    dims = [{"label": l, "val": int(v or 50), "color": _risk_color(v or 50)} for l, v in dim_defs]
    risk_level = {1: "低", 2: "中", 3: "高"}.get(profile.risk_level, _level_of(total)) if profile else _level_of(total)

    # 风险分历史：stat_daily 用户维度日行为序列（stat_aggregate 每日归档，近 30 天）
    # 分数为行为加权启发式（提交+40/点击+20/打开+10/举报-20，50 基准），仅趋势展示用
    labels: list[str] = []
    scores: list[int] = []
    trend_rows = db.execute(
        select(StatDaily.stat_date, StatDaily.open_cnt, StatDaily.click_cnt,
               StatDaily.submit_cnt, StatDaily.report_cnt)
        .where(StatDaily.dim_type == "user", StatDaily.dim_id == user_id)
        .order_by(StatDaily.stat_date)
        .limit(30)
    ).all()
    for d, o, c, s, r in trend_rows:
        labels.append(d.strftime("%m-%d"))
        scores.append(max(0, min(100, 50 + s * 40 + c * 20 + o * 10 - r * 20)))

    # 行为时间轴（近 20 条）
    action_map = {"open": "打开邮件", "click": "点击链接", "submit": "提交数据",
                  "attach_run": "运行附件", "report": "举报邮件", "bounce": "邮件退信"}
    t_rows = db.execute(
        select(TrackEvent, Campaign.name)
        .join(Campaign, Campaign.id == TrackEvent.campaign_id, isouter=True)
        .where(TrackEvent.user_id == user_id)
        .order_by(TrackEvent.created_at.desc())
        .limit(20)
    ).all()
    timeline = [
        {
            "time": e.created_at.strftime("%Y-%m-%d %H:%M") if e.created_at else "",
            "type": e.event_type,
            "title": c_name or f"演练#{e.campaign_id}",
            "desc": f"在演练「{c_name or e.campaign_id}」中{action_map.get(e.event_type, e.event_type)}",
        }
        for e, c_name in t_rows
    ]

    # 培训记录
    status_map = {"pending": "未开始", "learning": "学习中", "completed": "已完成", "overdue": "已超期"}
    a_rows = db.execute(
        select(TrainingAssignment, Course.title)
        .join(Course, Course.id == TrainingAssignment.course_id, isouter=True)
        .where(TrainingAssignment.user_id == user_id)
        .order_by(TrainingAssignment.assigned_at.desc())
    ).all()
    trainings = [
        {"name": c_title or f"课程#{a.course_id}", "progress": int(a.progress or 0),
         "status": status_map.get(a.status, a.status or ""),
         "completedAt": a.completed_at.strftime("%Y-%m-%d") if a.completed_at else ""}
        for a, c_title in a_rows
    ]

    return {
        "profile": {"dims": dims, "total": total, "riskLevel": risk_level},
        "trend": {"labels": labels, "scores": scores},
        "timeline": timeline,
        "trainings": trainings,
    }


# ---------- 导出（Excel / PDF） ----------

_SCOPES = {"campaign": "演练报表", "department": "部门对比报表",
           "trend": "综合趋势报表", "personal": "个人安全档案", "batch": "批量演练报表"}

_SECTION = dict  # {"title", "headers", "rows"}


def _campaign_sections(data: dict, campaign: Campaign) -> list[_SECTION]:
    """单演练报表导出分节：指标卡 → 漏斗 → 部门对比 → 中招明细。"""
    secs: list[_SECTION] = [
        {
            "title": "核心指标",
            "headers": ["指标", "数值", "说明"],
            "rows": [[m["title"], f"{m['value']}{m.get('suffix', '')}", m.get("sub", "")]
                     for m in data["metrics"]],
        },
        {
            "title": "转化漏斗",
            "headers": ["阶段", "数量", "转化率"],
            "rows": [[f["name"], f["value"], f"{f['rate']:.1f}%"] for f in data["funnel"]],
        },
        {
            "title": "部门对比",
            "headers": ["部门", "投递数", "中招数", "中招率%", "举报数", "举报率%"],
            "rows": [[d["dept"], d["sent"], d["victim"], f"{d['victimRate']:.1f}",
                      d["report"], f"{d['reportRate']:.1f}"] for d in data["deptCompare"]],
        },
        {
            "title": "中招/点击明细",
            "headers": ["姓名", "部门", "邮箱", "首次打开", "点击次数", "输入密码"],
            "rows": [[v["name"], v["dept"], v["email"], v["first_open"], v["clicks"],
                      "是" if v["input_pwd"] else "否"] for v in data["victims"]],
        },
    ]
    return secs


def _department_sections(data: dict, persons: dict | None = None) -> list[_SECTION]:
    """部门对比报表导出分节；persons 为 dept_persons_report 结果（人员明细）。"""
    secs: list[_SECTION] = [{
        "title": "部门横向对比",
        "headers": ["部门", "在职人数", "投递人次", "中招数", "打开率%", "点击率%",
                    "中招率%", "举报率%", "培训完成率%"],
        "rows": [[r["dept"], r["total"], r["targetCount"], r["victim"],
                  f"{r['openRate']:.1f}", f"{r['clickRate']:.1f}",
                  f"{r['submitRate']:.1f}", f"{r['reportRate']:.1f}",
                  f"{r['trainRate']:.1f}"] for r in data["rows"]],
    }]
    if persons and persons.get("rows"):
        secs.append({
            "title": "部门人员明细",
            "headers": ["姓名", "工号", "部门", "岗位", "参与次数", "中招率%", "风险"],
            "rows": [[p["name"], p["empNo"], p["dept"], p["position"], p["drills"],
                      f"{p['victimRate']:.1f}",
                      {"low": "低", "mid": "中", "high": "高"}.get(p["risk"], "低")]
                     for p in persons["rows"]],
        })
    return secs


def _trend_sections(data: dict) -> list[_SECTION]:
    secs: list[_SECTION] = [{
        "title": "月度趋势",
        "headers": ["月份", "演练次数", "中招率%", "举报率%"],
        "rows": [[data["labels"][i], data["campaignCounts"][i],
                  f"{data['submitRates'][i]:.1f}", f"{data['reportRates'][i]:.1f}"]
                 for i in range(len(data["labels"]))],
    }]
    if data["scenes"]:
        secs.append({
            "title": "场景维度",
            "headers": ["场景", "投递数", "中招率%", "举报率%"],
            "rows": [[s["scene"], s["targetCount"], f"{s['submitRate']:.1f}",
                      f"{s['reportRate']:.1f}"] for s in data["scenes"]],
        })
    return secs


def _personal_sections(data: dict, user: EmpUser) -> list[_SECTION]:
    dims = data["profile"]["dims"]
    secs: list[_SECTION] = [
        {
            "title": "五维风险画像",
            "headers": ["维度", "分值"],
            "rows": [[d["label"], d["val"]] for d in dims]
                    + [["综合风险值", data["profile"]["total"]],
                       ["风险等级", data["profile"]["riskLevel"]]],
        },
        {
            "title": "行为时间轴",
            "headers": ["时间", "事件", "说明"],
            "rows": [[t["time"], t["title"], t["desc"]] for t in data["timeline"]],
        },
    ]
    if data["trainings"]:
        secs.append({
            "title": "培训记录",
            "headers": ["课程", "完成日期", "进度%", "状态"],
            "rows": [[t["name"], t["completedAt"], t["progress"], t["status"]]
                     for t in data["trainings"]],
        })
    return secs


def _build_excel(title: str, sections: list[_SECTION]) -> bytes:
    """openpyxl 生成 xlsx：每节一行标题 + 表头 + 数据行。"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = (title or "报表")[:31]
    for sec in sections:
        ws.append([sec["title"]])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=12, color="378ADD")
        ws.append(sec["headers"])
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
        for row in sec["rows"]:
            ws.append(["" if v is None else str(v) for v in row])
        ws.append([])
    ws.freeze_panes = "A1"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_pdf(title: str, sections: list[_SECTION], operator: str) -> bytes:
    """WeasyPrint 渲染 HTML 表格：页脚带导出人/时间信息行（系统有 Noto Sans CJK 字体）。"""
    from html import escape
    from datetime import datetime
    from weasyprint import HTML

    body = ""
    for sec in sections:
        head = "".join(f"<th>{escape(str(h))}</th>" for h in sec["headers"])
        trs = "".join(
            "<tr>" + "".join(f"<td>{escape(str(v))}</td>" for v in row) + "</tr>"
            for row in sec["rows"]
        )
        body += f'<h3>{escape(sec["title"])}</h3><table><thead><tr>{head}</tr></thead><tbody>{trs}</tbody></table>'
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    html = f"""<!DOCTYPE html><html lang="zh"><head><meta charset="utf-8"><style>
      body {{ font-family: "Noto Sans CJK SC", "WenQuanYi Zen Hei", sans-serif; font-size: 12px; color: #333; }}
      h2 {{ color: #378ADD; margin: 0 0 4px; }}
      p.meta {{ color: #888; font-size: 11px; margin: 0 0 16px; }}
      h3 {{ margin: 18px 0 6px; color: #555; font-size: 13px; }}
      table {{ border-collapse: collapse; width: 100%; margin-bottom: 10px; }}
      th, td {{ border: 1px solid #d9d9d9; padding: 5px 8px; text-align: left; word-break: break-all; }}
      th {{ background: #f0f6ff; }}
      @page {{ size: A4 landscape; margin: 16mm;
        @bottom-center {{ content: "PhishLab 报表 · 导出人 {operator} · {now}"; font-size: 8px; color: #999; }} }}
    </style></head><body>
    <h2>{escape(title)}</h2>
    <p class="meta">导出人：{escape(operator)} ｜ 导出时间：{now} ｜ PhishLab 钓鱼演练平台</p>
    {body}
    </body></html>"""
    return HTML(string=html).write_pdf()


def export_report(db, account, kind: str, params: dict) -> tuple[bytes, str, str]:
    """同步生成导出文件（数据量小，直接生成返回流）。

    按 scope 复用现有报表查询拿数据 → openpyxl / WeasyPrint 生成 →
    返回 (文件字节, 文件名, media_type)；审计留痕含导出人与条件。
    """
    from datetime import datetime

    scope = params.get("scope", "campaign")
    range_ = params.get("range", "month")
    if kind not in ("excel", "pdf"):
        raise BizError(ErrorCode.PARAM_INVALID, "仅支持 excel/pdf 导出")
    if scope not in _SCOPES:
        raise BizError(ErrorCode.PARAM_INVALID, "导出范围不合法")
    if scope == "campaign" and not params.get("campaign_id"):
        raise BizError(ErrorCode.PARAM_INVALID, "导出演练报表需指定 campaign_id")
    if scope == "personal" and not params.get("user_id"):
        raise BizError(ErrorCode.PARAM_INVALID, "导出个人档案需指定 user_id")

    now = datetime.now()
    ext = "xlsx" if kind == "excel" else "pdf"
    filename = f"phishlab_{scope}_{now:%Y%m%d_%H%M%S}.{ext}"
    operator = getattr(account, "real_name", None) or account.username
    record_audit(db, account=account, module="report", action=f"export_{scope}_{kind}",
                 target_type=kind, detail=params or {})

    if scope == "campaign":
        cid = int(params["campaign_id"])
        campaign = db.get(Campaign, cid)
        if campaign is None:
            raise BizError(ErrorCode.NOT_FOUND)
        title = f"演练报表：{campaign.name}"
        sections = _campaign_sections(campaign_report(db, account, cid), campaign)
    elif scope == "department":
        title = f"部门对比报表（{range_}）"
        persons = None
        if params.get("dept_id"):
            persons = dept_persons_report(db, account, int(params["dept_id"]), range_)
            title = f"部门对比报表（{range_}）· 含人员明细"
        sections = _department_sections(department_report(db, account, range_), persons)
    elif scope == "trend":
        title = f"综合趋势报表（{range_}）"
        sections = _trend_sections(trend_report(db, account, range_))
    elif scope == "batch":
        cids = [int(c) for c in (params.get("campaign_ids") or [])]
        if not cids:
            raise BizError(ErrorCode.PARAM_INVALID, "批量导出需指定 campaign_ids")
        sections = []
        exported = 0
        for cid in cids:
            campaign = db.get(Campaign, cid)
            if campaign is None:
                continue
            sections += _campaign_sections(campaign_report(db, account, cid), campaign)
            exported += 1
        if not exported:
            raise BizError(ErrorCode.NOT_FOUND, "指定演练均不存在")
        title = f"批量演练报表（{exported} 场）"
    else:
        uid = int(params["user_id"])
        user = db.get(EmpUser, uid)
        if user is None:
            raise BizError(ErrorCode.NOT_FOUND)
        title = f"个人安全档案：{user.name}"
        sections = _personal_sections(personal_report(db, account, uid), user)

    if kind == "excel":
        return _build_excel(title, sections), filename, _XLSX_MIME
    return _build_pdf(title, sections, operator), filename, _PDF_MIME


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PDF_MIME = "application/pdf"
