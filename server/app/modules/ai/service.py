"""AI 服务：SSE 帧协议 {type: token|action|done|error}。

硬约束：AI 产出一律进 ai_draft 草稿态，人工 approve 后入库；审核记录入审计。
LLM Provider 适配层（app.modules.ai.llm）：配置了可用 Provider 时 Copilot/模板/分析
走真实 LLM（用量回填、留审计）；未配置或调用失败时优雅降级本地实现（演示环境保持可用）。
"""
import json
import logging
import re
from datetime import date, datetime

from sqlalchemy import case, func, select

from app.core.audit import record_audit
from app.core.errors import BizError, ErrorCode
from app.core.security import decrypt_secret, encrypt_secret

from .llm import get_client, get_provider, record_usage
from .models import AiDraft, AiMessage, AiProvider, AiSession

logger = logging.getLogger("phishlab.ai")

_SCENE_CN = {
    "finance": "财务报销", "hr": "HR通知", "system": "系统升级",
    "lottery": "中奖通知", "holiday": "节假日", "alert": "安全告警", "security": "安全告警",
}

_DEFAULT_SYSTEM_PROMPT = (
    "你是 PhishLab 企业网络安全钓鱼演练平台的 AI 助手，服务于安全运营人员。"
    "回答使用中文，简洁专业；涉及平台数据的回答基于用户可见数据，不编造演练结果；"
    "AI 产出的模板/文案属于草稿，需人工确认后生效。"
)

# 数据查询意图词表：命中即经 ChatBI 只读查询真实演练指标注入上下文
# （走红线 5 全链路守卫：只读账号 + 表白名单 + 结构化校验 + 数据权限注入 + 审计）
_DRILL_QUERY_WORDS = (
    "演练", "中招", "打开率", "点击率", "提交率", "上报", "举报", "投递",
    "退信", "趋势", "部门", "风险画像", "风险等级", "员工", "统计", "效果",
)
_MAX_INJECT_ROWS = 12  # 注入上下文的表格行数上限，控制 token 体积
# 意图 → ChatBI 稳定模板问法（原问题 LLM 自由生成失败时的兜底，模板问法不走 LLM 直查）
_DRILL_FALLBACK_QUERIES = (
    (re.compile(r"部门"), "各部门中招率"),
    (re.compile(r"趋势|走势"), "近7天中招趋势"),
    (re.compile(r"高危|风险"), "高危人员"),
    (re.compile(r"举报"), "举报最多的部门"),
    (re.compile(r"员工|最多|排行"), "中招最多的员工"),
    (re.compile(r"演练|效果|统计|分析"), "各演练统计"),
)


def _wants_drill_data(message: str) -> bool:
    """数据分析意图检测：命中任一数据性关键词即尝试注入（误触发由 ChatBI 兜底降级）。"""
    return any(w in message for w in _DRILL_QUERY_WORDS)


def _fallback_queries(message: str) -> list:
    """按意图关键词映射 ChatBI 模板问法（去重保序）。"""
    out = []
    for pattern, q in _DRILL_FALLBACK_QUERIES:
        if pattern.search(message) and q not in out:
            out.append(q)
    return out


def _rows_to_markdown(columns: list, rows: list) -> str:
    """查询结果转 markdown 表格，列多/行多截断控制注入体积。"""
    out = ["| " + " | ".join(str(c) for c in columns) + " |",
           "|" + "|".join("---" for _ in columns) + "|"]
    for r in rows[:_MAX_INJECT_ROWS]:
        out.append("| " + " | ".join(str(c) for c in r) + " |")
    if len(rows) > _MAX_INJECT_ROWS:
        out.append(f"\n（共 {len(rows)} 行，仅展示前 {_MAX_INJECT_ROWS} 行）")
    return "\n".join(out)


def _fmt_time(dt: datetime) -> str:
    seconds = max(int((datetime.now() - dt).total_seconds()), 0)
    if seconds < 60:
        return "刚刚"
    if seconds < 3600:
        return f"{seconds // 60}分钟前"
    if seconds < 86400:
        return f"{seconds // 3600}小时前"
    if seconds < 86400 * 30:
        return f"{seconds // 86400}天前"
    return dt.strftime("%Y-%m-%d")


def list_sessions(db, account):
    rows = db.scalars(
        select(AiSession).where(AiSession.account_id == account.id).order_by(AiSession.id.desc())
    ).all()
    return [
        {"id": s.id, "title": s.title or "新对话", "time": _fmt_time(s.created_at)}
        for s in rows
    ]


def list_messages(db, account, session_id: int):
    """会话消息明细（按时间正序）。仅本人会话可见——账号数据隔离，越权按不存在处理。"""
    session = db.get(AiSession, session_id)
    if session is None or session.account_id != account.id:
        raise BizError(ErrorCode.NOT_FOUND, "会话不存在")
    rows = db.scalars(
        select(AiMessage).where(AiMessage.session_id == session_id)
        .order_by(AiMessage.id.asc())
    ).all()
    return [
        {"id": m.id, "role": m.role, "content": m.content or "", "time": _fmt_time(m.created_at)}
        for m in rows
    ]


def _local_reply(db, message: str) -> str:
    """本地回复引擎：基于数据库真实统计的关键词应答（三期接 LLM）。"""
    from app.modules.campaign.models import Campaign, CampaignStat, CampaignTarget
    from app.modules.training.models import Course

    sent = CampaignTarget.send_status.in_(("sent", "delivered", "bounced", "failed"))
    total_submit = db.scalar(
        select(func.coalesce(func.sum(case((CampaignTarget.submit_flag == 1, 1), else_=0)), 0))
        .where(sent)
    ) or 0
    total_target = db.scalar(select(func.count()).where(sent)) or 0
    rate = total_submit / total_target * 100 if total_target else 0.0
    running = db.scalars(select(Campaign).where(Campaign.status == "running")).all()
    running_names = "、".join(c.name for c in running) or "暂无进行中的演练"
    courses = db.scalars(select(Course).order_by(Course.id).limit(3)).all()

    if any(k in message for k in ("演练", "分析", "效果", "中招")):
        return (
            f"根据当前平台数据：累计演练目标 {total_target:,} 人，中招 {total_submit:,} 人，"
            f"平均中招率 {rate:.1f}%。\n\n进行中的演练：{running_names}。"
            f"建议对中招率最高的部门优先安排「识别伪造发件人」专项培训。"
        )
    if any(k in message for k in ("模板", "钓鱼邮件", "话术")):
        return (
            "已为您梳理模板生成要点：\n1. 主题使用「紧急/通知/更新」等高频词可提升打开率约 40%；\n"
            "2. 建议在正文插入 {{.FirstName}}/{{.Department}} 动态变量提升可信度；\n"
            "3. 可通过「AI模板生成」页生成草稿，经人工确认后入库。"
        )
    if any(k in message for k in ("培训", "课程")):
        if courses:
            names = "、".join(c.title for c in courses)
            return f"推荐以下课程：{names}。可创建培训任务定向分配给中招人员，完成后自动结案。"
        return "当前课程库为空，可在「安全培训」页新建课程后再分配任务。"
    if any(k in message for k in ("举报", "报告")):
        return "举报数据可在「邮件举报」页查看：真实钓鱼举报需人工研判，判定后可推送 SOC 处置。"
    return (
        "我是您的安全演练 AI 助手，可以帮您：\n- 分析演练效果与中招趋势\n- 生成钓鱼邮件模板\n"
        "- 推荐培训课程\n\n*AI 生成内容处于草稿态，需人工确认后生效。审核记录将进入审计日志。*"
    )


def chat_stream(db, account, payload: dict):
    """Copilot 对话：返回 async generator 供 SSE。

    有可用 Provider → LLM 流式（历史上下文 + 用量回填 + 失败 error 帧）；
    未配置 Provider → 降级本地回复引擎（演示环境可用性兜底）。
    """
    message = (payload.get("message") or "").strip()
    if not message:
        raise BizError(ErrorCode.PARAM_INVALID, "消息不能为空")

    session = None
    if payload.get("session_id"):
        session = db.get(AiSession, payload["session_id"])
    if session is None:
        session = AiSession(account_id=account.id, title=message[:30],
                            page_context=payload.get("page_context"))
        db.add(session)
        db.flush()
    db.add(AiMessage(session_id=session.id, role="user", content=message))
    db.commit()

    provider = get_provider(db)
    if provider is None:
        # 降级：本地回复引擎（未配置 LLM Provider 时保持可用）
        reply = _local_reply(db, message)
        db.add(AiMessage(session_id=session.id, role="assistant", content=reply))
        db.commit()

        async def local_gen():
            for chunk in re.split(r"(?<=。)|(?<=\n)", reply):
                if chunk:
                    yield {"data": json.dumps({"type": "token", "content": chunk}, ensure_ascii=False)}
            yield {"data": json.dumps({"type": "done", "content": str(session.id)})}

        return local_gen()

    # LLM 路径：最近 20 条历史作为上下文（含本轮 user 消息）
    history = db.scalars(
        select(AiMessage).where(AiMessage.session_id == session.id)
        .order_by(AiMessage.id.desc()).limit(20)
    ).all()
    messages = [
        {"role": m.role, "content": (m.content or "")[:2000]}
        for m in reversed(history) if m.role in ("user", "assistant")
    ]
    system_prompt = provider.system_prompt or _DEFAULT_SYSTEM_PROMPT
    client = get_client(db, provider)

    async def llm_gen():
        full = ""
        tokens_in = tokens_out = None
        # 数据意图注入：ChatBI 只读查询真实指标后作为上下文给 LLM（失败降级，不阻断对话）。
        # 原问题优先（LLM 自由生成 SQL），失败按意图映射模板问法兜底，全失败则跳过注入。
        ctx_messages = list(messages)
        if _wants_drill_data(message):
            try:
                from . import chatbi as chatbi_svc
                result = None
                for q in [message, *_fallback_queries(message)]:
                    try:
                        result = await chatbi_svc.ask_question(db, account, q)
                        if result.get("rows"):
                            break
                    except BizError:
                        continue
                if result and result.get("rows"):
                    ctx_messages.append({
                        "role": "assistant",
                        "content": f"（以下为只读查询到的平台演练数据，回答用户时以真实数据为准，"
                                   f"与问题无关可忽略）\n{_rows_to_markdown(result['columns'], result['rows'])}",
                    })
                    logger.info("chatbi inject session=%s rows=%d", session.id, result["total"])
            except Exception:
                logger.info("chatbi inject skipped（查询失败不阻断对话）", exc_info=True)
        try:
            async for frame in client.stream(
                ctx_messages, system_prompt=system_prompt,
                temperature=float(provider.temperature or 0.7),
                max_tokens=provider.max_tokens or 2048,
            ):
                if frame["type"] == "token":
                    full += frame["content"]
                    yield {"data": json.dumps(
                        {"type": "token", "content": frame["content"]}, ensure_ascii=False)}
                elif frame["type"] == "usage":
                    tokens_in, tokens_out = frame.get("tokens_in"), frame.get("tokens_out")
            if not full.strip():
                raise BizError(ErrorCode.AI_PROVIDER_ERROR, "LLM 返回为空，请重试")
            db.add(AiMessage(session_id=session.id, role="assistant", content=full,
                             tokens_in=tokens_in, tokens_out=tokens_out))
            record_usage(db, provider, tokens_in, tokens_out)
            yield {"data": json.dumps({"type": "done", "content": str(session.id)})}
        except BizError as exc:
            logger.warning("llm stream biz error: %s", exc.message)
            yield {"data": json.dumps({"type": "error", "content": exc.message}, ensure_ascii=False)}
        except Exception:
            logger.exception("llm stream failed provider=%s", provider.id)
            yield {"data": json.dumps({"type": "error", "content": "AI 服务暂时不可用，请稍后重试"})}

    return llm_gen()


def _extract_json_meta(text: str) -> dict | None:
    """从 LLM 输出提取 JSON：直接解析 / ```json 代码块 / 首对花括号。失败返回 None。"""
    for candidate in (text.strip(),):
        try:
            data = json.loads(candidate)
            if isinstance(data, dict):
                return data
        except (json.JSONDecodeError, TypeError):
            pass
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.S)
    if m:
        try:
            data = json.loads(m.group(1))
            if isinstance(data, dict):
                return data
        except (json.JSONDecodeError, TypeError):
            pass
    m = re.search(r"\{.*\}", text, re.S)
    if m:
        try:
            data = json.loads(m.group(0))
            if isinstance(data, dict):
                return data
        except (json.JSONDecodeError, TypeError):
            pass
    return None


_TEMPLATE_SYSTEM = (
    "你是 PhishLab 钓鱼演练平台的邮件模板创作专家。根据用户要求生成一条演练邮件模板。"
    "只输出一个 JSON 对象，不要输出任何其他文字。JSON 字段："
    '{"name": 模板名, "subject": 邮件主题, "body": HTML 正文（可含 {{.FirstName}}/{{.ResetURL}} '
    '变量占位符，样式用内联 style）, "sender": 伪装发件人显示名, "difficulty": 1-3}。'
    "正文要可信、符合中文企业邮件习惯，不包含真实恶意内容。"
)


async def generate_template(db, account, payload: dict) -> int:
    """AI 模板生成 → ai_draft(biz_type=email_template)。

    LLM 可用时由 LLM 产出 JSON 草稿；未配置/调用失败/输出非法时降级本地模板。
    """
    scene = payload.get("scene", "finance")
    scene_cn = _SCENE_CN.get(scene, scene)
    audience = payload.get("audience") or "全体员工"
    tone = payload.get("tone") or "正式"
    difficulty = payload.get("difficulty", 2)

    meta = None
    fallback_reason = None
    provider = get_provider(db)
    if provider is not None:
        try:
            client = get_client(db, provider)
            result = await client.chat(
                [{"role": "user", "content":
                  f"场景：{scene_cn}；目标人群：{audience}；语气：{tone}；难度 {difficulty}。"
                  f"生成一条演练邮件模板。"}],
                system_prompt=provider.system_prompt or _TEMPLATE_SYSTEM,
                temperature=float(provider.temperature or 0.7),
                max_tokens=provider.max_tokens or 2048,
            )
            raw = _extract_json_meta(result["content"])
            if raw and raw.get("subject") and raw.get("body"):
                meta = {
                    "name": str(raw.get("name") or f"{scene_cn}·AI生成模板")[:64],
                    "scene": scene, "subject": str(raw["subject"])[:200],
                    "body": str(raw["body"]), "sender": str(raw.get("sender") or "信息安全部")[:64],
                    "difficulty": int(raw.get("difficulty") or difficulty),
                    "tone": tone, "audience": audience,
                }
                record_usage(db, provider, result.get("tokens_in"), result.get("tokens_out"))
            else:
                fallback_reason = "LLM 输出非预期 JSON，降级本地模板"
        except BizError as exc:
            fallback_reason = f"LLM 调用失败降级：{exc.message}"
        except Exception:
            logger.exception("generate_template llm failed")
            fallback_reason = "LLM 调用异常降级"

    if meta is None:
        subject = f"【{scene_cn}通知】请尽快处理您的待办事项"
        body = (
            f"<div style='font-family:sans-serif;line-height:1.8'>"
            f"<p>{{{{.FirstName}}}}，您好：</p>"
            f"<p>{scene_cn}相关事项需要您确认，请于今日 18:00 前点击下方链接完成操作。</p>"
            f"<p><a href='{{{{.ResetURL}}}}'>立即处理 →</a></p>"
            f"<p style='color:#888'>本邮件由系统自动发送（{tone}语气 · 目标人群：{audience}）</p></div>"
        )
        meta = {
            "name": f"{scene_cn}·AI生成模板", "scene": scene, "subject": subject,
            "body": body, "sender": "信息安全部", "difficulty": difficulty,
            "tone": tone, "audience": audience,
        }
    draft = AiDraft(
        biz_type="email_template",
        title=meta["subject"],
        content=json.dumps(meta, ensure_ascii=False),
        status="draft",
        created_by=account.id,
    )
    db.add(draft)
    db.commit()
    record_audit(db, account=account, module="ai", action="generate_template",
                 target_type="ai_draft", target_id=str(draft.id),
                 detail={"llm": provider is not None, "fallback": fallback_reason})
    return draft.id


_LANDING_SYSTEM = (
    "你是 PhishLab 钓鱼演练平台的仿冒登录页创作专家。根据用户要求生成一个演示用登录页面。"
    "只输出一个 JSON 对象，不要输出任何其他文字。JSON 字段："
    '{"name": 页面名称, "html_content": 完整 HTML 文档（<!DOCTYPE html> 开头，样式全用内联 style，'
    "移动端可加 meta viewport，不得引用外部脚本/样式/图片资源）, "
    '"fields": [{"label": 表单字段名, "input_type": "text|password", "sensitive_flag": 0|1}]}。'
    "页面仿照企业内部系统（邮箱/OA/网盘/认证门户）风格，含登录表单与提交按钮；"
    "口令字段用 input_type=password 且 sensitive_flag=1；不含任何真实恶意代码。"
)

_WECOM_SYSTEM = (
    "你是 PhishLab 钓鱼演练平台的企业微信消息模板创作专家。根据用户要求生成一条 textcard 卡片消息。"
    "只输出一个 JSON 对象，不要输出任何其他文字。JSON 字段："
    '{"name": 模板名, "title": 卡片标题（可含 {{.FirstName}}/{{.Department}} 变量）, '
    '"description": 卡片摘要（可含 {{.ResetURL}} 变量）, "btn_text": 按钮文案}。'
    "以内部部门名义（IT 部/HR/行政部）行文，严禁出现「微信安全中心/官方通知」等冒充官方字样。"
)

_ATTACH_SYSTEM = (
    "你是 PhishLab 钓鱼演练平台的诱饵文档创作专家。根据用户要求生成一份企业办公文档正文"
    "（如通知/明细/会议邀请/培训材料）。只输出一个 JSON 对象，不要输出任何其他文字。JSON 字段："
    '{"name": 文件名（不含扩展名）, "title": 文档标题, '
    '"paragraphs": ["正文段落1（可含 {{.FirstName}}/{{.Department}} 变量占位符）", "正文段落2"], '
    '"table": {"headers": ["列名"], "rows": [["单元格"]]} 或 null}。'
    "内容可信、符合中文企业文档习惯，不含真实恶意内容与外部链接。"
)

_LANDING_TYPE_MAP = {"mail": "mail_login", "oa": "oa_login", "pan": "pan_auth", "pay": "custom"}

_DEFAULT_LANDING_HTML = (
    "<!DOCTYPE html><html><head><meta charset=\"utf-8\">"
    "<meta name=\"viewport\" content=\"width=device-width,initial-scale=1\">"
    "<title>{{TITLE}}</title></head>"
    "<body style=\"font-family:Segoe UI,Arial,sans-serif;background:#f3f6fb;margin:0;padding:40px\">"
    "<div style=\"max-width:380px;margin:80px auto;background:#fff;border-radius:8px;"
    "padding:40px 32px;box-shadow:0 4px 20px rgba(0,0,0,.08)\">"
    "<div style=\"text-align:center;font-size:20px;font-weight:600;color:#0078d4;"
    "margin-bottom:8px\">🔒 {{TITLE}}</div>"
    "<div style=\"text-align:center;color:#666;font-size:13px;margin-bottom:28px\">请登录以继续</div>"
    "<form>"
    "<label style=\"display:block;font-size:13px;color:#333;margin-bottom:6px\">用户名 / 邮箱</label>"
    "<input type=\"text\" placeholder=\"请输入用户名\" "
    "style=\"width:100%;padding:10px 12px;border:1px solid #ddd;border-radius:4px;margin-bottom:16px;"
    "box-sizing:border-box\">"
    "<label style=\"display:block;font-size:13px;color:#333;margin-bottom:6px\">密码</label>"
    "<input type=\"password\" placeholder=\"请输入密码\" "
    "style=\"width:100%;padding:10px 12px;border:1px solid #ddd;border-radius:4px;margin-bottom:16px;"
    "box-sizing:border-box\">"
    "<button type=\"submit\" style=\"width:100%;background:#0078d4;color:#fff;border:none;padding:11px;"
    "border-radius:4px;font-size:15px;cursor:pointer\">登 录</button>"
    "</form><div style=\"text-align:center;margin-top:24px;font-size:11px;color:#999\">© 2026 {{TITLE}}</div>"
    "</div></body></html>"
)

_DEFAULT_LANDING_FIELDS = [
    {"label": "用户名", "input_type": "text", "sensitive_flag": 0, "sort": 0},
    {"label": "密码", "input_type": "password", "sensitive_flag": 1, "sort": 1},
    {"label": "验证码", "input_type": "text", "sensitive_flag": 0, "sort": 2},
]

_DEFAULT_EDU_TEXT = (
    "⚠️ 您刚刚中招了！\n\n这是一次公司组织的安全演练。您刚刚在仿冒页面输入了账号密码，"
    "如果在真实场景中，您的凭据已被攻击者窃取。\n\n请牢记：\n1. 认准官方域名，不轻信邮件中的链接\n"
    "2. 输入密码前核对网址是否为 HTTPS 且域名正确\n3. 可疑邮件请及时通过举报通道上报安全团队"
)


async def generate_landing(db, account, payload: dict) -> int:
    """AI 落地页生成 → ai_draft(biz_type=landing_page)。

    场景为视图类型（mail/oa/pan/…），映射到后端 type 枚举；LLM 不可用/输出非法时降级默认登录页。
    """
    scene = payload.get("scene") or "custom"
    page_type = _LANDING_TYPE_MAP.get(scene, scene)
    if page_type not in ("mail_login", "oa_login", "pan_auth", "custom"):
        page_type = "custom"
    company = (payload.get("company") or "企业内部系统").strip() or "企业内部系统"
    audience = payload.get("audience") or "全体员工"
    tone = payload.get("tone") or "正式"

    meta = None
    fallback_reason = None
    provider = get_provider(db)
    if provider is not None:
        try:
            client = get_client(db, provider)
            result = await client.chat(
                [{"role": "user", "content":
                  f"场景：{_SCENE_CN.get(scene, scene)}；企业：{company}；目标人群：{audience}；"
                  f"语气：{tone}。生成一个仿冒登录页面。"}],
                system_prompt=provider.system_prompt or _LANDING_SYSTEM,
                temperature=float(provider.temperature or 0.7),
                max_tokens=provider.max_tokens or 2048,
            )
            raw = _extract_json_meta(result["content"])
            if raw and raw.get("name") and raw.get("html_content"):
                fields = []
                for i, f in enumerate(raw.get("fields") or []):
                    if isinstance(f, dict) and f.get("label"):
                        fields.append({
                            "label": str(f["label"])[:32],
                            "input_type": f.get("input_type") or "text",
                            "sensitive_flag": 1 if f.get("sensitive_flag") else 0,
                            "sort": i,
                        })
                meta = {
                    "name": str(raw["name"])[:64],
                    "type": page_type,
                    "html_content": str(raw["html_content"]),
                    "form_schema": {"fields": fields or _DEFAULT_LANDING_FIELDS,
                                    "edu": _DEFAULT_EDU_TEXT, "redirect": "edu"},
                    "scene": scene, "tone": tone, "audience": audience,
                }
                record_usage(db, provider, result.get("tokens_in"), result.get("tokens_out"))
            else:
                fallback_reason = "LLM 输出非预期 JSON，降级本地默认页"
        except BizError as exc:
            fallback_reason = f"LLM 调用失败降级：{exc.message}"
        except Exception:
            logger.exception("generate_landing llm failed")
            fallback_reason = "LLM 调用异常降级"

    if meta is None:
        html = _DEFAULT_LANDING_HTML.replace("{{TITLE}}", company)
        meta = {
            "name": f"{company}登录页·AI生成",
            "type": page_type,
            "html_content": html,
            "form_schema": {"fields": _DEFAULT_LANDING_FIELDS,
                            "edu": _DEFAULT_EDU_TEXT, "redirect": "edu"},
            "scene": scene, "tone": tone, "audience": audience,
        }
    draft = AiDraft(
        biz_type="landing_page",
        title=meta["name"],
        content=json.dumps(meta, ensure_ascii=False),
        status="draft",
        created_by=account.id,
    )
    db.add(draft)
    db.commit()
    record_audit(db, account=account, module="ai", action="generate_landing",
                 target_type="ai_draft", target_id=str(draft.id),
                 detail={"llm": provider is not None, "fallback": fallback_reason})
    return draft.id


async def generate_wecom(db, account, payload: dict) -> int:
    """AI 企微消息模板生成 → ai_draft(biz_type=wecom_template)。

    产出 textcard（title/description/btn_text），审核入库时过合规红线校验。
    """
    scene = payload.get("scene", "system")
    scene_cn = _SCENE_CN.get(scene, scene)
    audience = payload.get("audience") or "全体员工"
    tone = payload.get("tone") or "正式"

    meta = None
    fallback_reason = None
    provider = get_provider(db)
    if provider is not None:
        try:
            client = get_client(db, provider)
            result = await client.chat(
                [{"role": "user", "content":
                  f"场景：{scene_cn}；目标人群：{audience}；语气：{tone}。生成一条企微 textcard 卡片消息。"}],
                system_prompt=provider.system_prompt or _WECOM_SYSTEM,
                temperature=float(provider.temperature or 0.7),
                max_tokens=provider.max_tokens or 2048,
            )
            raw = _extract_json_meta(result["content"])
            if raw and raw.get("title") and raw.get("description"):
                meta = {
                    "name": str(raw.get("name") or f"{scene_cn}·AI生成企微模板")[:64],
                    "msg_type": "textcard",
                    "title": str(raw["title"])[:128],
                    "description": str(raw["description"])[:512],
                    "btn_text": str(raw.get("btn_text") or "查看详情")[:16],
                    "url_mode": "track",
                    "scene": scene, "tone": tone, "audience": audience,
                }
                record_usage(db, provider, result.get("tokens_in"), result.get("tokens_out"))
            else:
                fallback_reason = "LLM 输出非预期 JSON，降级本地模板"
        except BizError as exc:
            fallback_reason = f"LLM 调用失败降级：{exc.message}"
        except Exception:
            logger.exception("generate_wecom llm failed")
            fallback_reason = "LLM 调用异常降级"

    if meta is None:
        meta = {
            "name": f"{scene_cn}·AI生成企微模板",
            "msg_type": "textcard",
            "title": f"【{scene_cn}】请及时确认您的信息",
            "description": f"{audience}您好：{scene_cn}相关事项需要您确认，"
                           f"请点击下方按钮查看详情。如有疑问请联系 IT 部。",
            "btn_text": "查看详情",
            "url_mode": "track",
            "scene": scene, "tone": tone, "audience": audience,
        }
    draft = AiDraft(
        biz_type="wecom_template",
        title=meta["title"],
        content=json.dumps(meta, ensure_ascii=False),
        status="draft",
        created_by=account.id,
    )
    db.add(draft)
    db.commit()
    record_audit(db, account=account, module="ai", action="generate_wecom",
                 target_type="ai_draft", target_id=str(draft.id),
                 detail={"llm": provider is not None, "fallback": fallback_reason})
    return draft.id


async def generate_attachment(db, account, payload: dict) -> int:
    """AI 诱饵文档生成 → ai_draft(biz_type=attachment)。

    仅良性文档（docx/xlsx）；宏/EXE 载荷属红线 6 默认关闭，不提供 AI 产出。
    确认入库时由 approve_draft 渲染真实文件写入附件库；
    docx 版可被投递链路注入 /pa/ beacon（附件运行追踪）并做占位符个性化。
    """
    scene = payload.get("scene", "通知")
    scene_cn = _SCENE_CN.get(scene, scene)
    audience = payload.get("audience") or "全体员工"
    tone = payload.get("tone") or "正式"
    doc_type = (payload.get("doc_type") or "docx").lower()
    if doc_type not in ("docx", "xlsx"):
        raise BizError(ErrorCode.PARAM_INVALID, "doc_type 仅支持 docx/xlsx（宏/EXE 载荷不开放）")

    meta = None
    fallback_reason = None
    provider = get_provider(db)
    if provider is not None:
        try:
            client = get_client(db, provider)
            result = await client.chat(
                [{"role": "user", "content":
                  f"场景：{scene_cn}；目标人群：{audience}；语气：{tone}；"
                  f"输出格式：{'Excel 表格（含 table）' if doc_type == 'xlsx' else 'Word 通知文档'}。"
                  f"生成一份诱饵文档正文。"}],
                system_prompt=provider.system_prompt or _ATTACH_SYSTEM,
                temperature=float(provider.temperature or 0.7),
                max_tokens=provider.max_tokens or 2048,
            )
            raw = _extract_json_meta(result["content"])
            paragraphs = [str(p) for p in (raw.get("paragraphs") or []) if p] if raw else []
            table = raw.get("table") if raw and isinstance(raw.get("table"), dict) else None
            if raw and raw.get("title") and (paragraphs or table):
                meta = {
                    "name": str(raw.get("name") or f"{scene_cn}通知")[:48],
                    "doc_type": doc_type,
                    "title": str(raw["title"])[:128],
                    "paragraphs": paragraphs[:30],
                    "table": table,
                    "scene": scene, "tone": tone, "audience": audience,
                }
                record_usage(db, provider, result.get("tokens_in"), result.get("tokens_out"))
            else:
                fallback_reason = "LLM 输出非预期 JSON，降级本地文档"
        except BizError as exc:
            fallback_reason = f"LLM 调用失败降级：{exc.message}"
        except Exception:
            logger.exception("generate_attachment llm failed")
            fallback_reason = "LLM 调用异常降级"

    if meta is None:
        doc_name = scene_cn if scene_cn.endswith(("通知", "明细", "邀请", "材料", "公告")) \
            else f"{scene_cn}通知"
        meta = {
            "name": doc_name,
            "doc_type": doc_type,
            "title": doc_name,
            "paragraphs": [
                "{{.FirstName}}，您好：",
                f"根据公司近期工作安排，现就「{scene_cn}」相关事项通知如下：",
                "1. 请于今日 18:00 前完成信息确认；",
                "2. 相关明细详见通知邮件正文，点击邮件内链接即可查看；",
                "3. 如有疑问，请联系本部门行政负责人。",
                "特此通知。",
            ],
            "table": {"headers": ["事项", "说明"],
                      "rows": [[f"{scene_cn}确认", "今日 18:00 前"],
                               ["咨询渠道", "本部门行政负责人"]]} if doc_type == "xlsx" else None,
            "scene": scene, "tone": tone, "audience": audience,
        }
    draft = AiDraft(
        biz_type="attachment",
        title=meta["title"],
        content=json.dumps(meta, ensure_ascii=False),
        status="draft",
        created_by=account.id,
    )
    db.add(draft)
    db.commit()
    record_audit(db, account=account, module="ai", action="generate_attachment",
                 target_type="ai_draft", target_id=str(draft.id),
                 detail={"doc_type": doc_type, "llm": provider is not None,
                         "fallback": fallback_reason})
    return draft.id


_ANALYSIS_SYSTEM = (
    "你是 PhishLab 安全演练平台的安全分析师。根据提供的平台真实指标数据撰写分析报告。"
    "只输出 Markdown 格式报告，包含：执行摘要、关键发现、改进建议。中文输出。"
)


async def generate_analysis(db, account, kind: str, target: dict) -> int:
    """智能分析报告（演练效果/部门画像/趋势预测/培训建议）→ ai_draft。

    LLM 可用时基于平台真实指标由 LLM 撰写；未配置/失败时降级本地生成。
    """
    from app.modules.campaign.models import Campaign, CampaignTarget

    sent = CampaignTarget.send_status.in_(("sent", "delivered", "bounced", "failed"))
    total_submit = db.scalar(
        select(func.coalesce(func.sum(case((CampaignTarget.submit_flag == 1, 1), else_=0)), 0))
        .where(sent)
    ) or 0
    total_target = db.scalar(select(func.count()).where(sent)) or 0
    rate = total_submit / total_target * 100 if total_target else 0.0
    campaign_cnt = db.scalar(select(func.count()).select_from(Campaign)) or 0

    kind_cn = {
        "campaign_effect": "演练效果分析", "dept_risk": "部门风险画像",
        "trend_forecast": "趋势预测", "training_recommend": "培训建议",
    }.get(kind, kind)

    content = None
    fallback_reason = None
    provider = get_provider(db)
    if provider is not None:
        metrics = (f"累计演练 {campaign_cnt} 场，覆盖 {total_target:,} 人次，"
                   f"平均中招率 {rate:.1f}%")
        try:
            client = get_client(db, provider)
            result = await client.chat(
                [{"role": "user", "content":
                  f"请撰写「{kind_cn}」报告。平台真实指标：{metrics}。"
                  f"目标参数：{json.dumps(target, ensure_ascii=False)}"}],
                system_prompt=provider.system_prompt or _ANALYSIS_SYSTEM,
                temperature=float(provider.temperature or 0.7),
                max_tokens=provider.max_tokens or 2048,
            )
            if result["content"].strip():
                content = result["content"]
                record_usage(db, provider, result.get("tokens_in"), result.get("tokens_out"))
            else:
                fallback_reason = "LLM 输出为空，降级本地生成"
        except BizError as exc:
            fallback_reason = f"LLM 调用失败降级：{exc.message}"
        except Exception:
            logger.exception("generate_analysis llm failed")
            fallback_reason = "LLM 调用异常降级"

    if content is None:
        content = (
            f"# {kind_cn}\n\n"
            f"## 执行摘要\n累计开展演练 {campaign_cnt} 场，覆盖 {total_target:,} 人次，"
            f"平均中招率 {rate:.1f}%。\n\n"
            f"## 关键发现\n- 财务类场景中招率显著高于其他场景\n"
            f"- 工作日 9:00-10:00 为打开高峰\n- 举报率与培训完成度呈正相关\n\n"
            f"## 改进建议\n1. 对高危部门开展定向专项培训\n2. 优化诱饵时间分布\n3. 完善举报反馈闭环"
        )
    draft = AiDraft(
        biz_type="report_summary",
        title=f"{kind_cn}（{datetime.now().strftime('%Y-%m-%d')}）",
        content=content,
        status="draft",
        created_by=account.id,
    )
    db.add(draft)
    db.commit()
    record_audit(db, account=account, module="ai", action="generate_analysis",
                 target_type="ai_draft", target_id=str(draft.id),
                 detail={"kind": kind, "llm": provider is not None, "fallback": fallback_reason})
    return draft.id


def list_drafts(db, account, status: str | None = None):
    stmt = select(AiDraft).where(AiDraft.status != "discarded").order_by(AiDraft.id.desc())
    if status:
        stmt = stmt.where(AiDraft.status == status)
    rows = db.scalars(stmt.limit(50)).all()
    from app.modules.account.models import SysAccount

    result = []
    for d in rows:
        reviewer = db.get(SysAccount, d.reviewer_id) if d.reviewer_id else None
        result.append({
            "id": d.id,
            "biz_type": d.biz_type,
            "title": d.title,
            "content": d.content,
            "status": d.status,
            "created_at": d.created_at.strftime("%Y-%m-%d %H:%M"),
            "reviewer": reviewer.real_name if reviewer else None,
            "reviewed_at": d.reviewed_at.strftime("%Y-%m-%d %H:%M") if d.reviewed_at else None,
        })
    return result


def _parse_draft_content(draft: AiDraft) -> dict:
    """草稿 content 解析为 dict，损坏内容按空对象处理（不阻断审核流）。"""
    try:
        meta = json.loads(draft.content or "{}")
        return meta if isinstance(meta, dict) else {}
    except (json.JSONDecodeError, TypeError):
        return {}


def _build_minimal_docx(meta: dict) -> bytes:
    """诱饵文档渲染为最小合法 docx（纯 zip 内 OOXML，无第三方依赖）。

    结构含 word/_rels/document.xml.rels（空关系集）——投递链路 _render_docx_variant
    依赖该文件追加 /pa/ beacon 外链关系；无该文件时 beacon 注入会被静默跳过。
    正文保留 {{.Xxx}} 占位符由投递链路按目标个性化。
    """
    import io
    import zipfile
    from xml.sax.saxutils import escape

    title = str(meta.get("title") or "通知")[:128]
    paragraphs = [str(p) for p in (meta.get("paragraphs") or []) if str(p).strip()]
    table = meta.get("table") if isinstance(meta.get("table"), dict) else None

    def _run(text: str, bold: bool = False) -> str:
        rpr = "<w:rPr><w:rFonts w:ascii=\"Calibri\" w:hAnsi=\"Calibri\" " \
              "w:eastAsia=\"宋体\"/><w:sz w:val=\"21\"/>" \
              + ("<w:b/><w:szCs w:val=\"21\"/>" if bold else "") + "</w:rPr>"
        return f'<w:r>{rpr}<w:t xml:space="preserve">{escape(str(text))}</w:t></w:r>'

    def _para(text: str, *, bold: bool = False, center: bool = False,
              size: int = 21, right: bool = False) -> str:
        jc = '<w:jc w:val="center"/>' if center else ('<w:jc w:val="right"/>' if right else "")
        ppr = f"<w:pPr>{jc}</w:pPr>" if jc else ""
        rpr = (f"<w:rPr><w:rFonts w:ascii=\"Calibri\" w:hAnsi=\"Calibri\" w:eastAsia=\"宋体\"/>"
               f"<w:sz w:val=\"{size}\"/><w:szCs w:val=\"{size}\"/>"
               + ("<w:b/>" if bold else "") + "</w:rPr>")
        return (f"<w:p>{ppr}<w:r>{rpr}"
                f"<w:t xml:space=\"preserve\">{escape(str(text))}</w:t></w:r></w:p>")

    body = [_para(title, bold=True, center=True, size=28), _para("")]
    body += [_para(p) for p in paragraphs]
    if table and (table.get("rows") or table.get("headers")):
        headers = [str(h) for h in (table.get("headers") or [])][:10] or ["内容"]
        grid = "".join(f'<w:gridCol w:w="{max(1200, 2400 // max(len(headers), 1))}"/>'
                       for _ in headers)
        trs = ["<w:tr>" + "".join(
            f'<w:tc><w:tcPr><w:tcW w:w="0" w:type="auto"/></w:tcPr>'
            f"<w:p>{_run(h, bold=True)}</w:p></w:tc>" for h in headers) + "</w:tr>"]
        for r in (table.get("rows") or [])[:60]:
            cells = (list(r) if isinstance(r, list) else [r])[:10]
            trs.append("<w:tr>" + "".join(
                f'<w:tc><w:tcPr><w:tcW w:w="0" w:type="auto"/></w:tcPr>'
                f"<w:p>{_run(c)}</w:p></w:tc>" for c in cells) + "</w:tr>")
        borders = "".join(
            f'<w:{side} w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
            for side in ("top", "left", "bottom", "right", "insideH", "insideV"))
        tbl = (f"<w:p/><w:tbl><w:tblPr><w:tblW w:w=\"0\" w:type=\"auto\"/>"
               f"<w:tblBorders>{borders}</w:tblBorders></w:tblPr>"
               f"<w:tblGrid>{grid}</w:tblGrid>{''.join(trs)}</w:tbl>")
        body.append(tbl)
    footer = " · ".join(filter(None, [str(meta.get("audience") or ""),
                                      f"{date.today().strftime('%Y年%m月%d日')} · IT 部"]))
    body += [_para(""), _para(footer, right=True)]
    body.append(
        '<w:sectPr><w:pgSz w:w="11906" w:h="16838"/>'
        '<w:pgMar w:top="1440" w:right="1440" w:bottom="1440" w:left="1440"/></w:sectPr>')

    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"<w:body>{''.join(body)}</w:body></w:document>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        "</Types>"
    )
    root_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="word/document.xml"/></Relationships>'
    )
    doc_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        "</Relationships>"
    )
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", content_types)
        z.writestr("_rels/.rels", root_rels)
        z.writestr("word/document.xml", document_xml)
        z.writestr("word/_rels/document.xml.rels", doc_rels)
    return out.getvalue()


def _render_ai_attachment(meta: dict) -> tuple[str, bytes]:
    """诱饵文档草稿渲染为真实文件：docx（自建 OOXML）/ xlsx（openpyxl）。

    返回 (文件名, 文件字节)。仅良性文档——宏/EXE 载荷属红线 6，AI 不产出。
    """
    import re

    doc_type = (meta.get("doc_type") or "docx").lower()
    paragraphs = [str(p) for p in (meta.get("paragraphs") or []) if str(p).strip()]
    table = meta.get("table") if isinstance(meta.get("table"), dict) else None
    if not paragraphs and not (table and (table.get("rows") or table.get("headers"))):
        raise BizError(ErrorCode.PARAM_INVALID, "文档内容为空，无法入库")
    name = re.sub(r"[\\/:*?\"<>|\s]+", "_", str(meta.get("name") or "AI文档"))[:48] or "AI文档"
    name = re.sub(r"\.(docx|xlsx)$", "", name, flags=re.I)
    title = str(meta.get("title") or "通知")[:128]

    if doc_type == "xlsx":
        import io
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "明细"
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        row = 3
        for p in paragraphs:
            ws.cell(row=row, column=1, value=p)
            row += 1
        if table:
            headers = [str(h) for h in (table.get("headers") or [])][:10] or ["内容"]
            row += 1
            for col, h in enumerate(headers, start=1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = Font(bold=True)
            row += 1
            for r in (table.get("rows") or [])[:60]:
                for col, v in enumerate((list(r) if isinstance(r, list) else [r])[:10], start=1):
                    ws.cell(row=row, column=col, value=str(v))
                row += 1
            for col, h in enumerate(headers, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(14, len(h) * 2 + 8)
        ws.column_dimensions["A"].width = 60
        ws.sheet_view.showGridLines = False
        buf = io.BytesIO()
        wb.save(buf)
        return f"{name}.xlsx", buf.getvalue()

    # docx（默认）：自建最小 OOXML（含 document.xml.rels，投递链路可注入 /pa/ beacon）
    return f"{name}.docx", _build_minimal_docx(meta)


def approve_draft(db, account, draft_id: int) -> dict:
    """确认入库：按 biz_type 写入目标表，回填 biz_id，记录审核人/时间。

    支持 email_template / landing_page / wecom_template / attachment；
    后三者经素材模块 service 层落库（含合规校验与审计），来源标记 ai。
    """
    from app.modules.template import service as tpl_service
    from app.modules.template.models import EmailTemplate

    draft = db.get(AiDraft, draft_id)
    if draft is None or draft.status == "discarded":
        raise BizError(ErrorCode.NOT_FOUND, "草稿不存在或已丢弃")
    if draft.status == "approved":
        raise BizError(ErrorCode.PARAM_INVALID, "草稿已确认入库")

    biz_id = None
    if draft.biz_type == "email_template":
        meta = _parse_draft_content(draft)
        tpl = EmailTemplate(
            name=meta.get("name") or (draft.title or "AI模板")[:128],
            scene=meta.get("scene") or "security",
            subject=meta.get("subject") or draft.title or "AI生成模板",
            html_body=meta.get("body") or draft.content or "",
            variables=sorted(set(re.findall(r"\{\{\.\w+\}\}", meta.get("body") or draft.content or ""))),
            source="ai",
            status="approved",
            created_by=account.id,
            sender=meta.get("sender"),
            stars=int(meta.get("difficulty") or 2),
        )
        db.add(tpl)
        db.flush()
        biz_id = tpl.id
    elif draft.biz_type == "landing_page":
        meta = _parse_draft_content(draft)
        if not (meta.get("name") and meta.get("html_content")):
            raise BizError(ErrorCode.PARAM_INVALID, "草稿缺少页面名称或 HTML 内容，无法入库")
        page_type = meta.get("type") or "custom"
        if page_type not in ("mail_login", "oa_login", "pan_auth", "custom"):
            page_type = "custom"
        biz_id = tpl_service.create_landing_page(db, account, {
            "name": meta["name"],
            "type": page_type,
            "html_content": meta["html_content"],
            "form_schema": meta.get("form_schema") or {"fields": []},
        }, source="ai")
    elif draft.biz_type == "wecom_template":
        meta = _parse_draft_content(draft)
        if not (meta.get("title") and meta.get("description")):
            raise BizError(ErrorCode.PARAM_INVALID, "草稿缺少卡片标题或摘要，无法入库")
        biz_id = tpl_service.create_wecom_template(db, account, {
            "name": meta.get("name") or (draft.title or "AI企微模板")[:128],
            "msg_type": "textcard",
            "title": str(meta["title"])[:128],
            "description": str(meta["description"])[:512],
            "btn_text": str(meta.get("btn_text") or "查看详情")[:16],
            "url_mode": "track",
        })
        # AI 草稿经人工确认入库，直接置为 approved（审核人/时间留痕）
        tpl_service.set_wecom_template_status(db, account, biz_id, "approved")
    elif draft.biz_type == "attachment":
        meta = _parse_draft_content(draft)
        filename, content = _render_ai_attachment(meta)
        biz_id = tpl_service.upload_attachment(db, account, filename, content, platform="AI生成")

    draft.status = "approved"
    draft.biz_id = biz_id
    draft.reviewer_id = account.id
    draft.reviewed_at = datetime.now()
    db.commit()
    record_audit(db, account=account, module="ai", action="approve_draft",
                 target_type="ai_draft", target_id=str(draft_id),
                 detail={"biz_type": draft.biz_type, "biz_id": biz_id})
    return {"id": draft_id, "status": "approved", "biz_id": biz_id}


def discard_draft(db, account, draft_id: int):
    draft = db.get(AiDraft, draft_id)
    if draft is None:
        raise BizError(ErrorCode.NOT_FOUND)
    draft.status = "discarded"
    draft.reviewer_id = account.id
    draft.reviewed_at = datetime.now()
    db.commit()
    record_audit(db, account=account, module="ai", action="discard_draft",
                 target_type="ai_draft", target_id=str(draft_id))
    return {"id": draft_id, "status": "discarded"}


# ---------- LLM Provider 管理（红线 2：Key 加密入库、API 只回显掩码） ----------

PROVIDER_TYPES = ("openai", "claude", "wenxin", "tongyi", "local")


def _mask_key(enc: bytes | None) -> str:
    """API Key 掩码回显：sk-****abcd（≤8 位全掩码；解密失败统一 ***）。"""
    if not enc:
        return ""
    try:
        key = decrypt_secret(enc)
    except Exception:
        return "***"
    if len(key) <= 8:
        return "*" * len(key)
    return key[:4] + "****" + key[-4:]


def _provider_usage(db, provider_ids: list[int]) -> dict[int, dict]:
    """最近 7 天用量聚合（provider_id → {calls, tokens_in, tokens_out}）。"""
    from datetime import timedelta

    from sqlalchemy import func as sa_func

    from .models import AiUsageStat

    if not provider_ids:
        return {}
    rows = db.execute(
        select(AiUsageStat.provider_id,
               sa_func.coalesce(sa_func.sum(AiUsageStat.call_count), 0),
               sa_func.coalesce(sa_func.sum(AiUsageStat.tokens_in), 0),
               sa_func.coalesce(sa_func.sum(AiUsageStat.tokens_out), 0))
        .where(AiUsageStat.provider_id.in_(provider_ids),
               AiUsageStat.stat_date >= (date.today() - timedelta(days=6)))
        .group_by(AiUsageStat.provider_id)
    ).all()
    return {pid: {"calls": int(c), "tokens_in": int(ti), "tokens_out": int(to)}
            for pid, c, ti, to in rows}


def list_providers(db) -> list[dict]:
    rows = db.scalars(select(AiProvider).order_by(AiProvider.id)).all()
    usage = _provider_usage(db, [p.id for p in rows])
    return [
        {
            "id": p.id, "name": p.name, "type": p.type, "endpoint": p.endpoint,
            "model": p.model, "api_key_masked": _mask_key(p.api_key_enc),
            "temperature": float(p.temperature or 0.7), "max_tokens": p.max_tokens or 2048,
            "system_prompt": p.system_prompt, "enabled": p.enabled == 1,
            "data_outbound": p.data_outbound == 1,
            "usage_7d": usage.get(p.id, {"calls": 0, "tokens_in": 0, "tokens_out": 0}),
        }
        for p in rows
    ]


def _check_provider_payload(payload: dict) -> None:
    ptype = payload.get("type")
    if ptype not in PROVIDER_TYPES:
        raise BizError(ErrorCode.PARAM_INVALID, f"type 仅支持 {PROVIDER_TYPES}")
    name = (payload.get("name") or "").strip()
    if not name or len(name) > 64:
        raise BizError(ErrorCode.PARAM_INVALID, "name 必填且 ≤64 字")
    endpoint = payload.get("endpoint") or ""
    if endpoint and not endpoint.startswith(("http://", "https://")):
        raise BizError(ErrorCode.PARAM_INVALID, "endpoint 需为 http(s):// 地址")
    temp = float(payload.get("temperature", 0.7))
    if not 0 <= temp <= 2:
        raise BizError(ErrorCode.PARAM_INVALID, "temperature 需在 0-2 之间")
    mt = int(payload.get("max_tokens", 2048))
    if not 1 <= mt <= 16384:
        raise BizError(ErrorCode.PARAM_INVALID, "max_tokens 需在 1-16384 之间")


def create_provider(db, account, payload: dict) -> dict:
    _check_provider_payload(payload)
    key = (payload.get("api_key") or "").strip()
    if not key and payload.get("type") != "local":
        raise BizError(ErrorCode.PARAM_INVALID, "api_key 必填（local 类型可留空）")
    p = AiProvider(
        name=payload["name"].strip(),
        type=payload["type"],
        endpoint=(payload.get("endpoint") or "").strip() or None,
        api_key_enc=encrypt_secret(key) if key else None,
        model=(payload.get("model") or "").strip() or None,
        temperature=payload.get("temperature", 0.7),
        max_tokens=payload.get("max_tokens", 2048),
        system_prompt=(payload.get("system_prompt") or "").strip() or None,
        enabled=1 if payload.get("enabled", True) else 0,
        data_outbound=1 if payload.get("data_outbound", True) else 0,
    )
    db.add(p)
    db.commit()
    record_audit(db, account=account, module="ai", action="provider_create",
                 target_type="ai_provider", target_id=str(p.id),
                 detail={"name": p.name, "type": p.type})
    return {"id": p.id, "api_key_masked": _mask_key(p.api_key_enc)}


def update_provider(db, account, provider_id: int, payload: dict) -> dict:
    p = db.get(AiProvider, provider_id)
    if p is None:
        raise BizError(ErrorCode.NOT_FOUND, "Provider 不存在")
    _check_provider_payload(payload)
    for field in ("name", "type", "endpoint", "model", "system_prompt"):
        if field in payload:
            setattr(p, field, (payload[field] or "").strip() or None)
    if "temperature" in payload:
        p.temperature = payload["temperature"]
    if "max_tokens" in payload:
        p.max_tokens = payload["max_tokens"]
    if "enabled" in payload:
        p.enabled = 1 if payload["enabled"] else 0
    if "data_outbound" in payload:
        p.data_outbound = 1 if payload["data_outbound"] else 0
    key = (payload.get("api_key") or "").strip()
    if key:
        p.api_key_enc = encrypt_secret(key)  # 更新 Key：旧密文覆盖，审计留痕
        record_audit(db, account=account, module="ai", action="provider_key_rotate",
                     target_type="ai_provider", target_id=str(provider_id),
                     detail={"name": p.name})
    db.commit()
    record_audit(db, account=account, module="ai", action="provider_update",
                 target_type="ai_provider", target_id=str(provider_id),
                 detail={"name": p.name, "type": p.type})
    return {"id": p.id, "api_key_masked": _mask_key(p.api_key_enc)}


def delete_provider(db, account, provider_id: int) -> None:
    p = db.get(AiProvider, provider_id)
    if p is None:
        raise BizError(ErrorCode.NOT_FOUND, "Provider 不存在")
    db.delete(p)
    db.commit()
    record_audit(db, account=account, module="ai", action="provider_delete",
                 target_type="ai_provider", target_id=str(provider_id),
                 detail={"name": p.name})


async def test_provider(db, provider_id: int) -> dict:
    """连通性测试：最小 chat 请求（16 token 上限），返回延迟与回复摘要。"""
    import time

    from .llm import get_client

    p = db.get(AiProvider, provider_id)
    if p is None:
        raise BizError(ErrorCode.NOT_FOUND, "Provider 不存在")
    client = get_client(db, p)
    started = time.monotonic()
    result = await client.chat([{"role": "user", "content": "回复 OK 即可"}],
                               system_prompt=p.system_prompt or "你是连通性测试助手，只回复 OK",
                               temperature=0.0, max_tokens=16)
    elapsed = round(time.monotonic() - started, 2)
    return {"ok": True, "latency_ms": int(elapsed * 1000),
            "reply": (result.get("content") or "")[:120],
            "tokens_in": result.get("tokens_in"), "tokens_out": result.get("tokens_out")}
